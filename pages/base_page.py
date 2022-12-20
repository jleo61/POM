from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium import webdriver
import pytest
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import logging
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from openpyxl import Workbook




class BasePage:
    def __init__(self, driver):
        self.driver = driver

    def click(self, locator):
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(locator)).click()

    def clear(self, locator):
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(locator)).clear()

    def clear_text_box(self, locator1):
        locator = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(locator1))
        locator.send_keys(Keys.CONTROL, "a")
        time.sleep(2)
        locator.send_keys(Keys.BACK_SPACE)
        time.sleep(2)

    def driver_title(self):
        return self.driver.title

        # NOT PARAMETERIZED
    def excel_table_data_scrapping_to_excel(self):
            num_col = self.driver.find_elements(By.XPATH, "//div[@class='left-align']/fieldset/table/tbody/tr[2]/td")
            num_row = self.driver.find_elements(By.XPATH, "//div[@class='left-align']/fieldset/table/tbody/tr")

            workbook = openpyxl.load_workbook("C:\\Users\\matne.LAPTOP-T1PULM73\\OneDrive\\Desktop\\page object model\\pages\\jj.xlsx")
            sheet = workbook.active
            for row in range(2, len(num_row) + 1):
                for col in range(1, len(num_col) + 1):
                    ele = self.driver.find_element(By.XPATH, "//div[@class='left-align']/fieldset/table/tbody/tr[" + str(row) + "]/td[" + str(col) + "]")

                    sheet.cell(row=row, column=col).value = ele.text
                workbook.save(filename="C:\\Users\\matne.LAPTOP-T1PULM73\\OneDrive\\Desktop\\page object model\\pages\\jj.xlsx")

    def drop_down_lists_get(self, locator):
        list_drop_down = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(locator))
        for ac in list_drop_down:
            print(ac.text)


    def excel_get_row_count(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        print(sheet.max_row)
        return sheet.max_row

    def excel_get_column_count(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        print(sheet.max_column)
        return sheet.max_column


    def logger(self, title):
        logger = logging.getLogger(title)
        fileHandler = logging.FileHandler('m43.log', mode='w')
        formatter = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s :%(message)s")
        formatter.datefmt="%Y-%m-%d %H:%M:%S"
        fileHandler.setFormatter(formatter)
        logger.addHandler(fileHandler)
        logger.setLevel(logging.INFO)

        return logger



    def send_keys(self, locator, value):
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(locator)).send_keys(value)

    # NOT PARAMETERIZED
    def select_list(self, locator):
        element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(locator))
        select = Select(element)
        select.select_by_visible_text("Option2")
        return select

    def get_attribute(self, locator):
        att_text = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(locator)).get_attribute("value")
        return att_text


    def wait_for_element(self, locator):
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(locator))

    def mouse_to_element(self, locator):
        mouse_action = ActionChains(self.driver)
        element = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(locator))

        mouse_action.move_to_element(element).perform()

    def alert_box_accept(self, locator):
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(locator)).click()
        alert_obj = Alert(self.driver)
        time.sleep(3)
        alert_obj.accept()

    def alert_box_dismiss(self, locator):
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(locator)).click()
        alert_obj = Alert(self.driver)
        time.sleep(3)
        alert_obj.dismiss()

    def scroll_window(self, value):
        self.driver.execute_script("window.scrollTo(0,"+value+")")

    # def mouse_action_dragdrop(self, source, destination):
    #     self.driver.find_element(source)
    #     self.driver.find_element(destination)
    #
    #     mouse_action = ActionChains(self.driver)
    #     mouse_action.drag_and_drop(source, destination)



    def add_table_column(self, locator):

        total = 0
        data = self.driver.find_elements(*locator)

        for i in data:
            total += int(i.text)
        return total

    def get_window_size(self):
        size = self.driver.get_window_size()
        print(size)




    def switch_to_iframe(self, locator):

        WebDriverWait(self.driver, 20).until(
            EC.frame_to_be_available_and_switch_to_it(locator))

    def switch_to_default_content(self):
        self.driver.switch_to.default_content()




    def get_len_count(self, locator):
        data = self.driver.find_elements(*locator)
        print(f"length is {data}")
        return data

    def save_screenshot(self, filepath):
        self.driver.save_screenshot(filepath)

    def get_cookies(self):
        cookies = self.driver.get_cookies()
        for ac in cookies:
            print(ac)
        return len(cookies)

    def delet_all_cookies(self):
        self.driver.delete_all_cookies()

    def delete_single_cookie(self, cookie_name):
        self.driver.delete_cookie(cookie_name)

    def date_picker(self, button_locator, element, next_button):
        self.click(button_locator)
        time.sleep(2)
        while True:
            ele = self.driver.find_element(*element)
            if ele.text == "March 2023":
                break
            else:
                self.click(next_button)

            # iterate through for a desired dates

    def link_text(self, linktext):
        self.driver.find_element(By.LINK_TEXT(linktext))

 # NOT PARAMETERIZED
    def simple_login_from_Excel(self):
        excel_path = "tests/user.xlsx"
        user_name_loc = By.XPATH, "//*[@id='app']/div[1]/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div[2]/input"
        password_loc = By.XPATH, "//*[@id='app']/div[1]/div/div[1]/div/div[2]/div[2]/form/div[2]/div/div[2]/input"
        log_in_button = By.XPATH, "//*[@id='app']/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button"

        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        for ac in range(2, sheet.max_row+1):

            user_name_input = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(user_name_loc))
            user_name_input.click()
            user_name_input_data = sheet.cell(row=ac, column=1).value
            user_name_input.send_keys(user_name_input_data)
            time.sleep(2)

            password_input = WebDriverWait(self.driver,10).until(EC.element_to_be_clickable(password_loc))
            password_input.click()
            password_input_data = sheet.cell(row=ac, column=2).value
            password_input.send_keys(password_input_data)
            time.sleep(2)

            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(log_in_button)).click()
            time.sleep(3)

            incorrect_input_displayed1 = self.driver.find_element(By.XPATH, "//*[@id='app']/div[1]/div/div[1]/div/div[2]/div[2]/div/div[1]/div[1]/p")

            if incorrect_input_displayed1:
                sheet.cell(row=ac, column=3, value="Pass")

            workbook.save(filename=excel_path)

    def back_space(self, locator):
        locator.send_keys(Keys.BACK_SPACE)


    def windowa_handle(self, current_handle, all_handles, locator):
        current_handle = self.driver.current_window_handle

        for ac in all_handles:
            print(ac)
            if ac != current_handle:
                self.driver.switch_to.window(ac)

                time.sleep(3)
                print(self.driver.title)
                time.sleep(2)

        # current_handle1 = self.driver.current_window_handle
        # print(current_handle1)
        #
        # # self.click(self.windo_switch_button)
        #
        # all_handles = self.driver.window_handles
        #
        # self.windowa_handle(current_handle1, all_handles, self.windo_switch_button)
        #
        # self.click(self.close_button)
        # self.click(self.course_button)
        # time.sleep(2)
        #
        # self.driver.close()
        #
        # self.driver.switch_to.window(current_handle1)










































































